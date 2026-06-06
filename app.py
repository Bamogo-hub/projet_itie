# app.py — Registre National des Bénéficiaires Effectifs
# Ministère de la Justice — République du Sénégal
# Flask + PostgreSQL (SQLAlchemy) — Compatible Python 3.11

import os, re, json, uuid, io, zipfile
from datetime import datetime, date
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, send_file, jsonify)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

# ── Application ───────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-me')

# ── Base de données ───────────────────────────────────────────────────────────
# Render fournit DATABASE_URL avec préfixe "postgres://" → corriger pour SQLAlchemy
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///beneficiaires_dev.db')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# ── Upload ────────────────────────────────────────────────────────────────────
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'zip', 'docx', 'xlsx', 'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── Modèles SQLAlchemy ────────────────────────────────────────────────────────
db = SQLAlchemy(app)

def gen_id():
    return str(uuid.uuid4())[:8].upper()

class User(db.Model):
    __tablename__ = 'users'
    id               = db.Column(db.String(8),   primary_key=True, default=gen_id)
    email            = db.Column(db.String(255),  unique=True, nullable=False)
    password_hash    = db.Column(db.String(512),  nullable=False)
    nom              = db.Column(db.String(100),  nullable=False)
    prenom           = db.Column(db.String(100),  nullable=False)
    role             = db.Column(db.String(20),   default='user')
    export_autorise  = db.Column(db.Boolean,      default=False)
    created_at       = db.Column(db.DateTime,     default=datetime.utcnow)
    declarations     = db.relationship('Declaration', backref='user', lazy=True,
                                       cascade='all, delete-orphan')

class Declaration(db.Model):
    __tablename__ = 'declarations'
    id               = db.Column(db.String(8),  primary_key=True, default=gen_id)
    date_saisie      = db.Column(db.DateTime,   default=datetime.utcnow)
    date_modification= db.Column(db.DateTime,   nullable=True)
    utilisateur_id   = db.Column(db.String(8),  db.ForeignKey('users.id'), nullable=False)
    statut           = db.Column(db.String(50), default='Soumis')
    # Section A
    denomination_sociale  = db.Column(db.String(300))
    ninea                 = db.Column(db.String(20),  unique=True)
    rccm                  = db.Column(db.String(100))
    forme_juridique       = db.Column(db.String(50))
    pays                  = db.Column(db.String(100))
    ville                 = db.Column(db.String(100))
    departement           = db.Column(db.String(100))
    siege_social          = db.Column(db.Text)
    greffe                = db.Column(db.String(100))
    telephone             = db.Column(db.String(30))
    email_entreprise      = db.Column(db.String(255))
    cotee_bourse          = db.Column(db.String(20), default='Non')
    filiale_cotee         = db.Column(db.String(20), default='Non')
    entreprise_etat       = db.Column(db.String(20), default='Non')
    autre_type_entite     = db.Column(db.String(20), default='Non')
    cotee_100             = db.Column(db.String(20), default='Non')
    nom_maison_mere       = db.Column(db.String(300))
    place_boursiere       = db.Column(db.String(200))
    numero_isin           = db.Column(db.String(50))
    pct_participation_mere= db.Column(db.Numeric(5,2))
    # Section B
    beneficiaire_nom      = db.Column(db.String(200))
    beneficiaire_prenom   = db.Column(db.String(200))
    sexe                  = db.Column(db.String(20))
    date_naissance        = db.Column(db.Date)
    lieu_naissance        = db.Column(db.String(200))
    nationalite           = db.Column(db.String(200))
    pays_residence        = db.Column(db.String(100))
    adresse_personnelle   = db.Column(db.Text)
    adresse_professionnelle=db.Column(db.Text)
    numero_cni            = db.Column(db.String(20))
    numero_passeport      = db.Column(db.String(30))
    telephone_ben         = db.Column(db.String(30))
    ppe                   = db.Column(db.String(20), default='Non')
    relation_ppe          = db.Column(db.String(20), default='Non')
    categorie_ppe         = db.Column(db.String(200))
    fonction_ppe          = db.Column(db.String(200))
    date_debut_fonction   = db.Column(db.Date)
    date_fin_fonction     = db.Column(db.Date)
    date_acquisition      = db.Column(db.Date)
    nature_relation_ppe   = db.Column(db.Text)
    # Section C
    parts_directes        = db.Column(db.String(20), default='Non')
    nombre_parts_directes = db.Column(db.Integer)
    pct_participation     = db.Column(db.Numeric(5,2))
    droits_vote_directs   = db.Column(db.String(20), default='Non')
    nombre_voix           = db.Column(db.Integer)
    pct_voix              = db.Column(db.Numeric(5,2))
    parts_indirectes      = db.Column(db.String(20), default='Non')
    nombre_parts_indirectes= db.Column(db.Integer)
    pct_participation_ind = db.Column(db.Numeric(5,2))
    representant_legal    = db.Column(db.String(20), default='Non')
    autres_beneficiaires  = db.Column(db.String(20), default='Non')
    date_beneficiaire     = db.Column(db.Date)
    intermediaires_json   = db.Column(db.Text)
    # Section D
    fichier_zip           = db.Column(db.String(500))
    fichiers_json         = db.Column(db.Text)
    certification_par     = db.Column(db.String(200))
    certification_fonction= db.Column(db.String(200))

    def to_dict(self):
        return {
            'id': self.id,
            'date_saisie': self.date_saisie.strftime('%d/%m/%Y %H:%M') if self.date_saisie else '',
            'date_saisie_iso': self.date_saisie.isoformat() if self.date_saisie else '',
            'utilisateur_id': self.utilisateur_id,
            'utilisateur_nom': f"{self.user.prenom} {self.user.nom}" if self.user else '',
            'utilisateur_email': self.user.email if self.user else '',
            'statut': self.statut or 'Soumis',
            'denomination_sociale': self.denomination_sociale or '',
            'ninea': self.ninea or '',
            'rccm': self.rccm or '',
            'forme_juridique': self.forme_juridique or '',
            'pays': self.pays or '',
            'ville': self.ville or '',
            'departement': self.departement or '',
            'siege_social': self.siege_social or '',
            'greffe': self.greffe or '',
            'telephone': self.telephone or '',
            'email_entreprise': self.email_entreprise or '',
            'cotee_bourse': self.cotee_bourse or 'Non',
            'filiale_cotee': self.filiale_cotee or 'Non',
            'entreprise_etat': self.entreprise_etat or 'Non',
            'autre_type_entite': self.autre_type_entite or 'Non',
            'nom_maison_mere': self.nom_maison_mere or '',
            'place_boursiere': self.place_boursiere or '',
            'numero_isin': self.numero_isin or '',
            'pourcentage_participation_mere': str(self.pct_participation_mere or ''),
            'beneficiaire_nom': self.beneficiaire_nom or '',
            'beneficiaire_prenom': self.beneficiaire_prenom or '',
            'sexe': self.sexe or '',
            'date_naissance': self.date_naissance.strftime('%d/%m/%Y') if self.date_naissance else '',
            'date_naissance_iso': self.date_naissance.isoformat() if self.date_naissance else '',
            'lieu_naissance': self.lieu_naissance or '',
            'nationalite': self.nationalite or '',
            'pays_residence': self.pays_residence or '',
            'adresse_personnelle': self.adresse_personnelle or '',
            'adresse_professionnelle': self.adresse_professionnelle or '',
            'numero_cni': self.numero_cni or '',
            'numero_passeport': self.numero_passeport or '',
            'telephone_ben': self.telephone_ben or '',
            'ppe': self.ppe or 'Non',
            'relation_ppe': self.relation_ppe or 'Non',
            'categorie_ppe': self.categorie_ppe or '',
            'fonction_ppe': self.fonction_ppe or '',
            'nature_relation_ppe': self.nature_relation_ppe or '',
            'parts_directes': self.parts_directes or 'Non',
            'nombre_parts_directes': self.nombre_parts_directes or '',
            'pourcentage_participation': str(self.pct_participation or ''),
            'droits_vote_directs': self.droits_vote_directs or 'Non',
            'nombre_voix': self.nombre_voix or '',
            'pourcentage_voix': str(self.pct_voix or ''),
            'parts_indirectes': self.parts_indirectes or 'Non',
            'nombre_parts_indirectes': self.nombre_parts_indirectes or '',
            'pourcentage_participation_indirecte': str(self.pct_participation_ind or ''),
            'representant_legal': self.representant_legal or 'Non',
            'autres_beneficiaires': self.autres_beneficiaires or 'Non',
            'date_beneficiaire': self.date_beneficiaire.strftime('%d/%m/%Y') if self.date_beneficiaire else '',
            'fichier_zip': self.fichier_zip or '',
            'fichiers': json.loads(self.fichiers_json) if self.fichiers_json else [],
            'certification_par': self.certification_par or '',
            'certification_fonction': self.certification_fonction or '',
        }

# ── Init DB automatique au démarrage ─────────────────────────────────────────
with app.app_context():
    db.create_all()
    if not User.query.first():
        db.session.add_all([
            User(id='ADMIN001', email='admin@justice.sn',
                 password_hash=generate_password_hash('Admin@2024'),
                 nom='Administrateur', prenom='Système',
                 role='admin', export_autorise=True),
            User(id='USER0001', email='user@justice.sn',
                 password_hash=generate_password_hash('User@2024'),
                 nom='Utilisateur', prenom='Test',
                 role='user', export_autorise=False),
        ])
        db.session.commit()
        print("✅ Base initialisée — comptes par défaut créés.")

# ── Référentiels ──────────────────────────────────────────────────────────────
REGIONS_SN = [
    {"region":"Dakar",        "departements":["Dakar","Pikine","Guédiawaye","Rufisque"]},
    {"region":"Thiès",        "departements":["Thiès","Mbour","Tivaouane"]},
    {"region":"Saint-Louis",  "departements":["Saint-Louis","Dagana","Podor"]},
    {"region":"Diourbel",     "departements":["Diourbel","Bambey","Mbacké"]},
    {"region":"Louga",        "departements":["Louga","Kébémer","Linguère"]},
    {"region":"Fatick",       "departements":["Fatick","Foundioune","Gossas"]},
    {"region":"Kaolack",      "departements":["Kaolack","Nioro du Rip","Guinguinéo"]},
    {"region":"Kaffrine",     "departements":["Kaffrine","Birkelane","Koungheul","Malem-Hodar"]},
    {"region":"Tambacounda",  "departements":["Tambacounda","Bakel","Goudiry","Koumpentoum"]},
    {"region":"Kédougou",     "departements":["Kédougou","Salémata","Saraya"]},
    {"region":"Kolda",        "departements":["Kolda","Médina Yoro Fula","Vélingara"]},
    {"region":"Sédhiou",      "departements":["Sédhiou","Bounkiling","Goudomp"]},
    {"region":"Ziguinchor",   "departements":["Ziguinchor","Bignona","Oussouye"]},
    {"region":"Matam",        "departements":["Matam","Kanel","Ranérou-Ferlo"]},
    {"region":"Bignona",      "departements":["Bignona"]},
]

ALL_EXPORT_COLUMNS = [
    ('id','ID'), ('date_saisie','Date Saisie'), ('utilisateur_nom','Saisi par'),
    ('denomination_sociale','Dénomination Sociale'), ('ninea','NINEA'), ('rccm','RCCM'),
    ('forme_juridique','Forme Juridique'), ('pays','Pays'), ('ville','Région'),
    ('departement','Département'), ('siege_social','Siège Social'),
    ('telephone','Téléphone'), ('email_entreprise','Email'),
    ('cotee_bourse','Cotée Bourse'), ('entreprise_etat',"Entreprise d'État"),
    ('beneficiaire_nom','Nom'), ('beneficiaire_prenom','Prénom'),
    ('sexe','Sexe'), ('date_naissance','Date Naissance'),
    ('nationalite','Nationalité'), ('pays_residence','Pays Résidence'),
    ('numero_cni','N° CNI'), ('numero_passeport','Passeport'),
    ('ppe','PPE'), ('categorie_ppe','Catégorie PPE'), ('fonction_ppe','Fonction PPE'),
    ('parts_directes','Parts Directes'), ('pourcentage_participation','% Part.'),
    ('droits_vote_directs','Droits Vote'), ('pourcentage_voix','% Voix'),
    ('parts_indirectes','Parts Indirectes'), ('pourcentage_participation_indirecte','% Indirect'),
    ('representant_legal','Représentant Légal'), ('statut','Statut'),
]

@app.context_processor
def inject_globals():
    return {
        'now': datetime.now,
        'regions_senegal': REGIONS_SN,
        'formes_juridiques': ['SA','SARL','SUARL','SAS','SNC','SCS','GIE',
                               'Association','ONG','Coopérative','EP','SEM','Autre'],
    }

# ── Helpers ───────────────────────────────────────────────────────────────────
def allowed_file(f):
    return '.' in f and f.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def parse_date(s):
    if not s or str(s).strip() == '': return None
    try: return datetime.strptime(str(s).strip(), '%Y-%m-%d').date()
    except: return None

def parse_decimal(s):
    if not s or str(s).strip() == '': return None
    try: return float(s)
    except: return None

def parse_int(s):
    if not s or str(s).strip() == '': return None
    try: return int(float(s))
    except: return None

# ── Validation ────────────────────────────────────────────────────────────────
def ok_ninea(v):
    if not v: return True
    return bool(re.match(r'^\d{10}[A-Z]{1,2}\d?$', v.upper().strip()))

def ok_cni(v):
    if not v: return True
    return bool(re.match(r'^\d{13}$', v.strip()))

def ok_phone(v):
    if not v: return True
    c = re.sub(r'[\s\-\(\)]','',v.strip())
    return bool(re.match(r'^(\+|00)\d{7,15}$', c))

def ok_email(v):
    if not v: return True
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]{2,}$', v.strip()))

def ok_pct(v):
    if not v or str(v).strip() == '': return True, None
    try:
        f = float(v)
        if f < 2:   return False, "Minimum 2%."
        if f > 100: return False, "Maximum 100%."
        return True, None
    except: return False, "Valeur numérique invalide."

# ── Décorateurs ───────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def d(*a,**kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a,**kw)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a,**kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash("Accès refusé.",'error'); return redirect(url_for('dashboard'))
        return f(*a,**kw)
    return d

def export_required(f):
    @wraps(f)
    def d(*a,**kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        u = db.session.get(User, session['user_id'])
        if not u or not u.export_autorise:
            flash("Export non autorisé. Contactez l'administrateur.",'error')
            return redirect(url_for('liste_declarations'))
        return f(*a,**kw)
    return d

# ════════════════════════════════════════════════════════════════
#  ROUTES
# ════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email','').strip()
        pwd   = request.form.get('password','')
        u = User.query.filter_by(email=email).first()
        if u and check_password_hash(u.password_hash, pwd):
            session.update({'user_id':u.id,'email':u.email,'nom':u.nom,
                            'prenom':u.prenom,'role':u.role,
                            'export_autorise':u.export_autorise})
            flash(f'Bienvenue, {u.prenom} {u.nom} !','success')
            return redirect(url_for('dashboard'))
        flash('Email ou mot de passe incorrect.','error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear(); flash('Vous avez été déconnecté.','info')
    return redirect(url_for('login'))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    q = Declaration.query
    if session.get('role') != 'admin':
        q = q.filter_by(utilisateur_id=session['user_id'])
    decls = q.all()
    today = date.today()
    stats = {
        'total':        len(decls),
        'aujourd_hui':  sum(1 for d in decls if d.date_saisie and d.date_saisie.date()==today),
        'entreprises':  len(set(d.ninea for d in decls if d.ninea)),
        'beneficiaires':len(set(d.beneficiaire_nom for d in decls if d.beneficiaire_nom)),
        'ppe':          sum(1 for d in decls if d.ppe=='Oui'),
    }
    monthly = {}
    for d in decls:
        m = d.date_saisie.strftime('%Y-%m') if d.date_saisie else ''
        if m: monthly[m] = monthly.get(m,0)+1
    sm = sorted(monthly.items())[-12:]
    formes = {}
    for d in decls:
        f = d.forme_juridique or 'Non renseigné'; formes[f]=formes.get(f,0)+1
    recent = sorted(decls, key=lambda x: x.date_saisie or datetime.min, reverse=True)[:5]
    return render_template('dashboard.html', stats=stats,
        monthly_labels=json.dumps([m[0] for m in sm]),
        monthly_data=json.dumps([m[1] for m in sm]),
        formes_labels=json.dumps(list(formes.keys())),
        formes_data=json.dumps(list(formes.values())),
        recent=recent)

# ── Nouvelle déclaration ──────────────────────────────────────────────────────
@app.route('/declaration/nouvelle', methods=['GET','POST'])
@login_required
def nouvelle_declaration():
    if request.method == 'POST':
        form   = request.form
        errors = []
        ninea  = form.get('ninea','').strip().upper()
        cni    = form.get('numero_cni','').strip()
        tel    = form.get('telephone','').strip()
        email  = form.get('email_entreprise','').strip()

        if not ninea:                errors.append("Le NINEA est obligatoire.")
        elif not ok_ninea(ninea):    errors.append(f"NINEA invalide « {ninea} » — ex: 0043385142U2")
        elif Declaration.query.filter_by(ninea=ninea).first():
            errors.append(f"NINEA {ninea} déjà enregistré.")
        if cni   and not ok_cni(cni):     errors.append("N° CNI invalide — 13 chiffres requis.")
        if tel   and not ok_phone(tel):   errors.append("Numéro de téléphone invalide.")
        if email and not ok_email(email): errors.append("Email entreprise invalide.")
        for fname,flabel in [('pourcentage_participation','% Participation directe'),
                              ('pourcentage_voix','% Droits de vote'),
                              ('pourcentage_participation_indirecte','% Participation indirecte'),
                              ('pourcentage_participation_mere','% Participation maison mère')]:
            ok,msg = ok_pct(form.get(fname,''))
            if not ok: errors.append(f"{flabel} : {msg}")

        if errors:
            for e in errors: flash(e,'error')
            return render_template('declaration.html', form_data=form)

        # Fichiers
        fichier_zip, fichiers_lst = '', []
        if 'documents' in request.files:
            files = request.files.getlist('documents')
            valid = [f for f in files if f and f.filename and allowed_file(f.filename)]
            if valid:
                folder = os.path.join(app.config['UPLOAD_FOLDER'],
                                      f'Ent_{ninea}_{datetime.now().strftime("%Y%m%d%H%M%S")}')
                os.makedirs(folder, exist_ok=True)
                zip_path = os.path.join(folder, 'documents.zip')
                with zipfile.ZipFile(zip_path,'w') as zf:
                    for file in valid:
                        fn = secure_filename(file.filename)
                        fp = os.path.join(folder, fn)
                        file.save(fp); zf.write(fp, fn)
                        fichiers_lst.append({'nom':fn,'path':fp,'type':fn.rsplit('.',1)[-1].lower()})
                fichier_zip = zip_path

        intermediaires = []
        for i in range(3):
            n = form.get(f'intermediaire_nom_{i}','').strip()
            if n: intermediaires.append({'nom':n,'id':form.get(f'intermediaire_id_{i}','')})

        d = Declaration(
            utilisateur_id=session['user_id'],
            ninea=ninea,
            denomination_sociale=form.get('denomination_sociale',''),
            rccm=form.get('rccm','').upper(),
            forme_juridique=form.get('forme_juridique',''),
            pays=form.get('pays',''), ville=form.get('ville',''),
            departement=form.get('departement',''),
            siege_social=form.get('siege_social',''),
            greffe=form.get('greffe',''),
            telephone=tel, email_entreprise=email,
            cotee_bourse=form.get('cotee_bourse','Non'),
            filiale_cotee=form.get('filiale_cotee','Non'),
            entreprise_etat=form.get('entreprise_etat','Non'),
            autre_type_entite=form.get('autre_type_entite','Non'),
            cotee_100=form.get('cotee_100','Non'),
            nom_maison_mere=form.get('nom_maison_mere',''),
            place_boursiere=form.get('place_boursiere',''),
            numero_isin=form.get('numero_isin','').upper(),
            pct_participation_mere=parse_decimal(form.get('pourcentage_participation_mere')),
            beneficiaire_nom=form.get('beneficiaire_nom','').upper(),
            beneficiaire_prenom=form.get('beneficiaire_prenom',''),
            sexe=form.get('sexe',''),
            date_naissance=parse_date(form.get('date_naissance')),
            lieu_naissance=form.get('lieu_naissance',''),
            nationalite=form.get('nationalite',''),
            pays_residence=form.get('pays_residence',''),
            adresse_personnelle=form.get('adresse_personnelle',''),
            adresse_professionnelle=form.get('adresse_professionnelle',''),
            numero_cni=cni,
            numero_passeport=form.get('numero_passeport','').upper(),
            telephone_ben=form.get('telephone_ben',''),
            ppe=form.get('ppe','Non'), relation_ppe=form.get('relation_ppe','Non'),
            categorie_ppe=form.get('categorie_ppe',''),
            fonction_ppe=form.get('fonction_ppe',''),
            date_debut_fonction=parse_date(form.get('date_debut_fonction')),
            date_fin_fonction=parse_date(form.get('date_fin_fonction')),
            date_acquisition=parse_date(form.get('date_acquisition_propriete')),
            nature_relation_ppe=form.get('nature_relation_ppe',''),
            parts_directes=form.get('parts_directes','Non'),
            nombre_parts_directes=parse_int(form.get('nombre_parts_directes')),
            pct_participation=parse_decimal(form.get('pourcentage_participation')),
            droits_vote_directs=form.get('droits_vote_directs','Non'),
            nombre_voix=parse_int(form.get('nombre_voix')),
            pct_voix=parse_decimal(form.get('pourcentage_voix')),
            parts_indirectes=form.get('parts_indirectes','Non'),
            nombre_parts_indirectes=parse_int(form.get('nombre_parts_indirectes')),
            pct_participation_ind=parse_decimal(form.get('pourcentage_participation_indirecte')),
            representant_legal=form.get('representant_legal','Non'),
            autres_beneficiaires=form.get('autres_beneficiaires','Non'),
            date_beneficiaire=parse_date(form.get('date_beneficiaire')),
            intermediaires_json=json.dumps(intermediaires),
            fichier_zip=fichier_zip,
            fichiers_json=json.dumps(fichiers_lst),
            certification_par=form.get('certification_par',''),
            certification_fonction=form.get('certification_fonction',''),
        )
        db.session.add(d); db.session.commit()
        flash(f'Déclaration #{d.id} soumise avec succès !','success')
        return redirect(url_for('liste_declarations'))
    return render_template('declaration.html', form_data={})

# ── Liste ─────────────────────────────────────────────────────────────────────
@app.route('/declarations')
@login_required
def liste_declarations():
    q = Declaration.query
    if session.get('role') != 'admin':
        q = q.filter_by(utilisateur_id=session['user_id'])
    search=request.args.get('search','')
    filtre_ppe=request.args.get('ppe','')
    filtre_forme=request.args.get('forme','')
    if search:
        like=f'%{search}%'
        q=q.filter(db.or_(Declaration.ninea.ilike(like),Declaration.rccm.ilike(like),
                           Declaration.denomination_sociale.ilike(like),
                           Declaration.beneficiaire_nom.ilike(like)))
    if filtre_ppe:   q=q.filter_by(ppe=filtre_ppe)
    if filtre_forme: q=q.filter_by(forme_juridique=filtre_forme)
    decls = q.order_by(Declaration.date_saisie.desc()).all()
    u = db.session.get(User, session['user_id'])
    export_ok = u.export_autorise if u else False
    return render_template('liste.html', declarations=decls,
        search=search, filtre_ppe=filtre_ppe, filtre_forme=filtre_forme,
        export_ok=export_ok)

# ── Détail ────────────────────────────────────────────────────────────────────
@app.route('/declaration/<decl_id>')
@login_required
def detail_declaration(decl_id):
    d = db.session.get(Declaration, decl_id)
    if not d: flash('Introuvable.','error'); return redirect(url_for('liste_declarations'))
    if session.get('role')!='admin' and d.utilisateur_id!=session['user_id']:
        flash('Accès refusé.','error'); return redirect(url_for('liste_declarations'))
    return render_template('detail.html', declaration=d.to_dict(), decl=d)

# ── Modifier ──────────────────────────────────────────────────────────────────
@app.route('/declaration/<decl_id>/modifier', methods=['GET','POST'])
@login_required
def modifier_declaration(decl_id):
    d = db.session.get(Declaration, decl_id)
    if not d: flash('Introuvable.','error'); return redirect(url_for('liste_declarations'))
    if session.get('role')!='admin' and d.utilisateur_id!=session['user_id']:
        flash('Accès refusé.','error'); return redirect(url_for('liste_declarations'))
    if request.method == 'POST':
        form=request.form; errors=[]
        cni=form.get('numero_cni','').strip()
        ninea=form.get('ninea','').strip().upper()
        tel=form.get('telephone','').strip()
        if cni   and not ok_cni(cni):     errors.append('CNI invalide.')
        if ninea and not ok_ninea(ninea):  errors.append('NINEA invalide.')
        if tel   and not ok_phone(tel):    errors.append('Téléphone invalide.')
        for fn,fl in [('pourcentage_participation','% Participation'),
                      ('pourcentage_voix','% Voix'),('pourcentage_participation_indirecte','% Indirect')]:
            ok,msg=ok_pct(form.get(fn,'')); 
            if not ok: errors.append(f'{fl} : {msg}')
        if errors:
            for e in errors: flash(e,'error')
            return render_template('declaration.html',form_data=form,edit_mode=True,decl_id=decl_id)
        d.denomination_sociale=form.get('denomination_sociale','')
        d.ninea=ninea; d.rccm=form.get('rccm','').upper()
        d.forme_juridique=form.get('forme_juridique','')
        d.pays=form.get('pays',''); d.ville=form.get('ville','')
        d.departement=form.get('departement','')
        d.siege_social=form.get('siege_social','')
        d.telephone=tel; d.email_entreprise=form.get('email_entreprise','')
        d.beneficiaire_nom=form.get('beneficiaire_nom','').upper()
        d.beneficiaire_prenom=form.get('beneficiaire_prenom','')
        d.sexe=form.get('sexe','')
        d.date_naissance=parse_date(form.get('date_naissance'))
        d.nationalite=form.get('nationalite','')
        d.pays_residence=form.get('pays_residence','')
        d.adresse_personnelle=form.get('adresse_personnelle','')
        d.adresse_professionnelle=form.get('adresse_professionnelle','')
        d.numero_cni=cni; d.numero_passeport=form.get('numero_passeport','').upper()
        d.ppe=form.get('ppe','Non'); d.categorie_ppe=form.get('categorie_ppe','')
        d.fonction_ppe=form.get('fonction_ppe','')
        d.parts_directes=form.get('parts_directes','Non')
        d.nombre_parts_directes=parse_int(form.get('nombre_parts_directes'))
        d.pct_participation=parse_decimal(form.get('pourcentage_participation'))
        d.droits_vote_directs=form.get('droits_vote_directs','Non')
        d.pct_voix=parse_decimal(form.get('pourcentage_voix'))
        d.parts_indirectes=form.get('parts_indirectes','Non')
        d.pct_participation_ind=parse_decimal(form.get('pourcentage_participation_indirecte'))
        d.representant_legal=form.get('representant_legal','Non')
        d.certification_par=form.get('certification_par','')
        d.certification_fonction=form.get('certification_fonction','')
        d.date_modification=datetime.utcnow()
        db.session.commit()
        flash('Déclaration modifiée !','success')
        return redirect(url_for('detail_declaration', decl_id=decl_id))
    return render_template('declaration.html', form_data=d.to_dict(), edit_mode=True, decl_id=decl_id)

# ── Supprimer ─────────────────────────────────────────────────────────────────
@app.route('/declaration/<decl_id>/supprimer', methods=['POST'])
@admin_required
def supprimer_declaration(decl_id):
    d = db.session.get(Declaration, decl_id)
    if d: db.session.delete(d); db.session.commit(); flash('Déclaration supprimée.','success')
    return redirect(url_for('liste_declarations'))

# ── Documents ─────────────────────────────────────────────────────────────────
@app.route('/declaration/<decl_id>/documents')
@login_required
def voir_documents(decl_id):
    d = db.session.get(Declaration, decl_id)
    if not d: flash('Introuvable.','error'); return redirect(url_for('liste_declarations'))
    if session.get('role')!='admin' and d.utilisateur_id!=session['user_id']:
        flash('Accès refusé.','error'); return redirect(url_for('liste_declarations'))
    fichiers = json.loads(d.fichiers_json) if d.fichiers_json else []
    return render_template('documents.html', declaration=d.to_dict(), fichiers=fichiers)

@app.route('/declaration/<decl_id>/fichier/<filename>')
@login_required
def servir_fichier(decl_id, filename):
    d = db.session.get(Declaration, decl_id)
    if not d: return "Introuvable", 404
    if session.get('role')!='admin' and d.utilisateur_id!=session['user_id']:
        return "Accès refusé", 403
    fichiers = json.loads(d.fichiers_json) if d.fichiers_json else []
    f = next((x for x in fichiers if x['nom']==filename), None)
    if not f or not os.path.exists(f['path']): return "Fichier introuvable", 404
    ext = filename.rsplit('.',1)[-1].lower()
    mimes = {'pdf':'application/pdf','png':'image/png','jpg':'image/jpeg','jpeg':'image/jpeg',
             'docx':'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
             'xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','zip':'application/zip'}
    return send_file(f['path'], mimetype=mimes.get(ext,'application/octet-stream'),
                     as_attachment=ext not in ('pdf','png','jpg','jpeg'), download_name=filename)

# ── Export ────────────────────────────────────────────────────────────────────
@app.route('/export/choix')
@login_required
@export_required
def export_choix():
    return render_template('export_choix.html', all_columns=ALL_EXPORT_COLUMNS)

@app.route('/export/excel', methods=['GET','POST'])
@login_required
@export_required
def export_excel():
    q = Declaration.query
    if session.get('role')!='admin': q=q.filter_by(utilisateur_id=session['user_id'])
    decls = q.all()
    selected = (request.form.getlist('colonnes') if request.method=='POST'
                else request.args.getlist('colonnes'))
    if not selected: selected = [k for k,_ in ALL_EXPORT_COLUMNS]
    col_map = {k:l for k,l in ALL_EXPORT_COLUMNS if k in selected}
    ordered = [(k,col_map[k]) for k,_ in ALL_EXPORT_COLUMNS if k in col_map]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Bénéficiaires Effectifs"
    hfill=PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
    hfont=Font(color="FFFFFF",bold=True,size=11)
    afill=PatternFill(start_color="EBF3FB",end_color="EBF3FB",fill_type="solid")
    for col,(key,label) in enumerate(ordered,1):
        c=ws.cell(row=1,column=col,value=label)
        c.fill=hfill; c.font=hfont; c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=28
    for ri,d in enumerate(decls,2):
        dd=d.to_dict()
        for col,(key,_) in enumerate(ordered,1):
            ws.cell(row=ri,column=col,value=str(dd.get(key,'') or ''))
        if ri%2==0:
            for cell in ws[ri]: cell.fill=afill
    for col in ws.columns:
        mx=max((len(str(c.value or '')) for c in col),default=8)
        ws.column_dimensions[col[0].column_letter].width=min(mx+3,42)
    ws2=wb.create_sheet("Infos"); ws2.append(["Généré le",datetime.now().strftime('%d/%m/%Y %H:%M')])
    ws2.append(["Déclarations",len(decls)]); ws2.append(["Par",f"{session.get('prenom','')} {session.get('nom','')}"])
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'beneficiaires_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

# ── Admin ─────────────────────────────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin():
    return render_template('admin.html', users=User.query.order_by(User.created_at).all())

@app.route('/admin/utilisateur/nouveau', methods=['POST'])
@admin_required
def nouveau_utilisateur():
    email=request.form.get('email','').strip()
    if User.query.filter_by(email=email).first():
        flash('Email déjà utilisé.','error'); return redirect(url_for('admin'))
    u=User(email=email, password_hash=generate_password_hash(request.form.get('password','')),
           nom=request.form.get('nom',''), prenom=request.form.get('prenom',''),
           role=request.form.get('role','user'),
           export_autorise=request.form.get('role','user')=='admin')
    db.session.add(u); db.session.commit()
    flash(f'Utilisateur {u.prenom} {u.nom} créé.','success')
    return redirect(url_for('admin'))

@app.route('/admin/utilisateur/<user_id>/export', methods=['POST'])
@admin_required
def toggle_export(user_id):
    u=db.session.get(User,user_id)
    if u:
        u.export_autorise=not u.export_autorise; db.session.commit()
        flash(f"Export {'autorisé' if u.export_autorise else 'révoqué'} pour {u.prenom} {u.nom}.","success")
    return redirect(url_for('admin'))

@app.route('/admin/utilisateur/<user_id>/supprimer', methods=['POST'])
@admin_required
def supprimer_utilisateur(user_id):
    if user_id==session['user_id']:
        flash("Impossible de supprimer votre propre compte.",'error')
        return redirect(url_for('admin'))
    u=db.session.get(User,user_id)
    if u: db.session.delete(u); db.session.commit(); flash('Utilisateur supprimé.','success')
    return redirect(url_for('admin'))

# ── API Validation ────────────────────────────────────────────────────────────
@app.route('/api/valider/ninea')
def api_ninea():
    v=request.args.get('v',''); ok=ok_ninea(v)
    return jsonify({'valide':ok,'message':'' if ok else 'Format invalide. Ex: 0043385142U2'})

@app.route('/api/valider/cni')
def api_cni():
    v=request.args.get('v',''); ok=ok_cni(v)
    return jsonify({'valide':ok,'message':'' if ok else '13 chiffres exacts requis.'})

@app.route('/api/valider/telephone')
def api_telephone():
    v=request.args.get('v',''); ok=ok_phone(v)
    return jsonify({'valide':ok,'message':'' if ok else 'Numéro invalide.'})

if __name__ == '__main__':
    app.run(debug=False, port=5000)
